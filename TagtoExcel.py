import tkinter
from tkinter import *
from tkinter.scrolledtext import ScrolledText

import os
import time

import openpyxl
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.worksheet.table import Table, TableStyleInfo

import mutagen
from mutagen.mp3 import MP3
from mutagen.easyid3 import EasyID3 as MP3
from mutagen.easyid3 import EasyID3

top = tkinter.Tk()
top.geometry("310x450")
top.title("TagtoExcel")
top.resizable(False, False)

def createExcel():
    textPad.insert(END, 'Creating Excel File...\n')
    path = os.path.dirname(os.path.realpath(__file__))

    files = []

    for r, d, f in os.walk(path):
        for file in f:
            if ".mp3" in file:
                files.append(file)


    wb = openpyxl.Workbook()
    sheet = wb.active

    sheet["A1"] = "Song File Name"
    sheet.column_dimensions["A"].width = 36
    currentCell = sheet['A1']
    currentCell.alignment = Alignment(horizontal='center')

    sheet["B1"] = "Title"
    sheet.column_dimensions["B"].width = 36
    currentCell = sheet['B1']
    currentCell.alignment = Alignment(horizontal='center')

    sheet["C1"] = "Album/Movie"
    sheet.column_dimensions["C"].width = 36
    currentCell = sheet['C1']
    currentCell.alignment = Alignment(horizontal='center')

    sheet["D1"] = "Year"
    sheet.column_dimensions["D"].width = 10
    currentCell = sheet['D1']
    currentCell.alignment = Alignment(horizontal='center')

    sheet["E1"] = "Genre"
    sheet.column_dimensions["E"].width = 15
    currentCell = sheet['E1']
    currentCell.alignment = Alignment(horizontal='center')
    
    rowNum = 2
    done = True

    for file in files:
        m = MP3(file)
        
        currentCell = sheet.cell(row = rowNum, column=1).value = file

        currentCell = sheet.cell(row = rowNum, column=2).value = m["title"][0]

        currentCell = sheet.cell(row = rowNum, column=3).value = m["album"][0]
        
        dateArray = m.get("date", ["-1"])
        if dateArray[0] == "-1":
            date = "No Year"
        else:
            try:
                date = int(dateArray[0])
            except ValueError:
                textPad.insert(END, files[rowNum-2]+ "\n" + 'has an unacceptable date\n')
                done = False
                break
                
        currentCell = sheet.cell(row = rowNum, column=4).value = date


        genreArray = m.get("genre", ["-1"])
        if genreArray[0] == "-1":
            genre = "No Genre"
        else:
            genre = genreArray[0]
        currentCell = sheet.cell(row = rowNum, column=5).value = genre
        rowNum += 1

    tab = Table(displayName="Table1", ref="A1:E"+str(len(files)+1))
    style = TableStyleInfo(name="TableStyleMedium14", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    sheet.add_table(tab)

    if done:
        wb.save('songs.xlsx')
        textPad.insert(END, 'Excel File "Songs.xslx" created\n')
    else:
        textPad.insert(END, 'Excel File NOT created\n')

def readExcelTitle():
    textPad.insert(END, 'Writing Title...\n')
    path = "songs.xlsx"
    wb = openpyxl.load_workbook(path)

    sheet = wb.active
    max_row = sheet.max_row

    names = []
    title = []

    for i in range(2, max_row +1):
        info = sheet.cell(row=i, column=1)
        names.append(info.value)

    for i in range(2, max_row +1):
        info = sheet.cell(row=i, column=2)
        title.append(info.value)


    for i in range(0, len(names)):
        try:
            tag = EasyID3(names[i])
            tag["title"] = title[i]
            tag.save(v2_version=3)
        except:
            tag = mutagen.File(path, easy=True)
            tag.add_tags()

    textPad.insert(END, 'Titles Correctly Written\n')
    
def readExcelAlbum():
    textPad.insert(END, 'Writing Album Name...\n')
    path = "songs.xlsx"
    wb = openpyxl.load_workbook(path)

    sheet = wb.active
    max_row = sheet.max_row

    names = []
    album = []

    for i in range(2, max_row +1):
        info = sheet.cell(row=i, column=1)
        names.append(info.value)

    for i in range(2, max_row +1):
        info = sheet.cell(row=i, column=3)
        album.append(info.value)


    for i in range(0, len(names)):
        try:
            tag = EasyID3(names[i])
            tag["album"] = album[i]
            tag.save(v2_version=3)
        except:
            tag = mutagen.File(path, easy=True)
            tag.add_tags()

    textPad.insert(END, 'Album Names Correctly Written\n')

def readExcelYear():
    textPad.insert(END, 'Writing Year...\n')
    path = "songs.xlsx"
    wb = openpyxl.load_workbook(path)

    sheet = wb.active
    max_row = sheet.max_row

    names = []
    date = []

    for i in range(2, max_row +1):
        info = sheet.cell(row=i, column=1)
        names.append(info.value)

    for i in range(2, max_row +1):
        info = sheet.cell(row=i, column=4)
        date.append(info.value)


    for i in range(0, len(names)):
        try:
            tag = EasyID3(names[i])
            tag["date"] = str(date[i])
            tag.save(v2_version=3)
        except:
            tag = mutagen.File(path, easy=True)
            tag.add_tags()

    textPad.insert(END, 'Years Correctly Written\n')

def readExcelGenre():
    textPad.insert(END, 'Writing Genres...\n')
    path = "songs.xlsx"
    wb = openpyxl.load_workbook(path)

    sheet = wb.active
    max_row = sheet.max_row

    names = []
    genre = []

    for i in range(2, max_row +1):
        info = sheet.cell(row=i, column=1)
        names.append(info.value)

    for i in range(2, max_row +1):
        info = sheet.cell(row=i, column=5)
        genre.append(info.value)


    for i in range(0, len(names)):
        try:
            tag = EasyID3(names[i])
            tag["genre"] = genre[i]
            tag.save(v2_version=3)
        except:
            tag = mutagen.File(path, easy=True)
            tag.add_tags()

    textPad.insert(END, 'Genres Correctly Written\n')

lbl_title = Label(top, text="Written By Deep Harquissandas")
emp1 = Label(top, text="")
createExcel = Button(top, text="Create Excel File", command=createExcel, width=40)
emp2 = Label(top, text="")
writeTitle = Button(top, text="Write Title From Excel File", command=readExcelTitle, width=40)
emp3 = Label(top, text="")
writeAlbum = Button(top, text="Write Album From Excel File", command=readExcelAlbum, width=40)
emp4 = Label(top, text="")
writeYear = Button(top, text="Write Year From Excel File", command=readExcelYear, width=40)
emp5 = Label(top, text="")
writeGenre = Button(top, text="Write Genre From Excel File", command=readExcelGenre, width=40)
emp6 = Label(top, text="")
textPad = ScrolledText(top, width=33, height=10)

lbl_title.pack()
emp1.pack()
createExcel.pack()
emp2.pack()
writeTitle.pack()
emp3.pack()
writeAlbum.pack()
emp4.pack()
writeYear.pack()
emp5.pack()
writeGenre.pack()
emp6.pack()
textPad.pack()


top.mainloop()
